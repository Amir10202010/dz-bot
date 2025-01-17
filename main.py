import os
import logging
import datetime
from openpyxl import load_workbook
from openpyxl.cell import MergedCell
from datetime import datetime, timedelta, date

days_of_week = {
    0: 'Понедельник',
    1: 'Вторник',
    2: 'Среду',
    3: 'Четверг',
    4: 'Пятницу',
    5: 'Суббота',
    6: 'Воскресенье'
}

months = {
    1: 'Января', 2: 'Февраля', 3: 'Марта', 4: 'Апреля', 5: 'Мая', 6: 'Июня',
    7: 'Июля', 8: 'Августа', 9: 'Сентября', 10: 'Октября', 11: 'Ноября', 12: 'Декабря'
}

async def parser_merged_cell(sheet, cell):
    if isinstance(cell, MergedCell):
        for merged_range in sheet.merged_cells.ranges:
            if cell.coordinate in merged_range:
                cell = sheet.cell(row=merged_range.min_row, column=merged_range.min_col)
                break
    return cell.value


async def edit_merged_cell(sheet, cell, sheet2):
    if isinstance(cell, MergedCell):
        for merged_range in sheet.merged_cells.ranges:
            if cell.coordinate in merged_range:
                cell = sheet2.cell(row=merged_range.min_row, column=merged_range.min_col)
                break
    return cell


async def dz_day(day, week):
    today = datetime.now()
    current_weekday = today.weekday()
    days_until_monday = (7 - current_weekday) % 7 
    days_to_target = days_until_monday + (week * 7) + day 
    target_date = today + timedelta(days=days_to_target)
    day_of_week = days_of_week[target_date.weekday()]
    month = months[target_date.month]
    return f"{day_of_week} {target_date.day} {month} {target_date.year} года"


async def get_week_number():
    start_date = date(2024, 9, 15)
    current_date = date.today()
    delta = current_date - start_date
    week_number = int(delta.days / 7) + 1
    filename = f"dz.xlsx"
    wb = load_workbook(filename)

    if f"week-{week_number-2}" in wb.sheetnames:
        wb.remove_sheet(wb[f'week-{week_number-2}'])

    if os.path.exists(f"week-{week_number-2}.jpg"):
        os.remove(f"week-{week_number}.jpg")

    return week_number


async def process_excel_file(number):
    week_number = await get_week_number() + number
    week_sheet_name = f"week-{week_number}"
    filename = f"dz.xlsx"

    wb = load_workbook(filename)
    source = wb["example"]

    if week_sheet_name not in wb.sheetnames:
        wb.copy_worksheet(source).title = week_sheet_name
        wb.save(filename)


async def dz_for_day(day, number, group):
    day = (day + 1) * 2
    filename = f"dz.xlsx"
    week_number = await get_week_number() + number
    dz = []
    await process_excel_file(number)
    wb = load_workbook(filename)
    ws = wb[f'week-{week_number}']

    if group == 1:
        dz.append('для 1 группы! 📚')
        for cell in ws[day - 1]:
            value = await parser_merged_cell(ws, cell)
            if value is not None:
                dz.append(str(value).replace('\n', '  '))
    else:
        dz.append('для 2 группы! 📚')
        for cell in ws[day]:
            value = await parser_merged_cell(ws, cell)
            if value is not None:
                dz.append(str(value).replace('\n', '  ')) 

    return '\n'.join(list(dict.fromkeys(dz)))


async def edit_dz(day, number, group, subject, text):
    day = (day + 1) * 2
    filename = f"dz.xlsx"
    week_number = await get_week_number() + number
    await process_excel_file(number)
    wb = load_workbook(filename)
    ws = wb['example']
    ws2 = wb[f"week-{week_number}"]

    if group == 1:
        for cell in ws[day - 1]:
            value = await parser_merged_cell(ws, cell)
            if str(value) == subject:
                new_cell = await edit_merged_cell(ws, cell, ws2)
                ws2.cell(row=new_cell.row, column=new_cell.column).value = f"{value}\n{text}"
    else:
        for cell in ws[day]:
            value = await parser_merged_cell(ws, cell)
            if str(value) == subject:
                new_cell = await edit_merged_cell(ws, cell, ws2)
                ws2.cell(row=new_cell.row, column=new_cell.column).value = f"{value}\n{text}"

    wb.save(filename)


async def subjects_for_day(day, number, group):
    day = (day + 1) * 2
    filename = f"dz.xlsx"
    dz = []
    await process_excel_file(number)
    wb = load_workbook(filename)
    ws = wb[f'example']

    if group == 1:
        for cell in ws[day - 1]:
            value = await parser_merged_cell(ws, cell)
            if value is not None:
                dz.append(str(value))
    else:
        for cell in ws [day]:
            value = await parser_merged_cell(ws, cell)
            if value is not None:
                dz.append(str(value)) 

    return list(dict.fromkeys(dz))
